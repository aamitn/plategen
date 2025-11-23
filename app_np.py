# NAMEPLATE LIST EXCEL/PDF GENERATOR 
import sqlite3
import os
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QComboBox, QCheckBox, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QHBoxLayout, QMessageBox, QHeaderView
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import QSpinBox
import sys
import urllib.request
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PyQt6.QtWidgets import QFileDialog, QMessageBox

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


DB_FILE = 'nameplates.db'
DB_URL = 'https://gitlab.com/aamitn/assets/-/raw/main/nameplate_excel_liveline/nameplates.db'
DB_SCHEMA_PY = 'app_np_db_schema.py'
DB_SCHEMA_EXE = 'app_np_db_schema.exe'

def ensure_database():
    # 1️⃣ If DB already exists, nothing to do
    if os.path.exists(DB_FILE):
        print(f"Database '{DB_FILE}' already exists.")
        return True

    # 2️⃣ Try to download DB
    print(f"Database '{DB_FILE}' not found. Trying to download from remote...")
    try:
        urllib.request.urlretrieve(DB_URL, DB_FILE)
        print(f"Database downloaded successfully from {DB_URL}")
        return True
    except Exception as e:
        print(f"Failed to download database: {e}")
        print("Will attempt to create database locally...")

    # 3️⃣ Run schema only if download failed
    if os.path.exists(DB_SCHEMA_PY):
        print(f"Running schema script '{DB_SCHEMA_PY}'...")
        try:
            subprocess.run([sys.executable, DB_SCHEMA_PY], check=True)
            print("Database created successfully via Python schema.")
            return True
        except subprocess.CalledProcessError as ex:
            print(f"Failed to run {DB_SCHEMA_PY}: {ex}")

    if os.path.exists(DB_SCHEMA_EXE):
        print(f"Running schema executable '{DB_SCHEMA_EXE}'...")
        try:
            subprocess.run([DB_SCHEMA_EXE], check=True)
            print("Database created successfully via EXE schema.")
            return True
        except subprocess.CalledProcessError as ex:
            print(f"Failed to run {DB_SCHEMA_EXE}: {ex}")

    print("All attempts to create the database failed.")
    return False


# Ensure DB is ready
if not ensure_database():
    raise RuntimeError("Database setup failed. Cannot continue.")

# ------------------- DB Functions -------------------
def fetch_nameplates(ch_group_id):
    """Fetch nameplates for a given charger group ID, including COMMON entries with repeater handling."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM ch_groups WHERE group_name='COMMON'")
    common_id = cursor.fetchone()[0]

    repeater_map = {
        'SFCB': [''],               
        'DFCB': ['(Charger 1)', '(Charger 2)'],
        'FFCB': ['(FC)', '(FCB)']
    }

    cursor.execute("SELECT group_name FROM ch_groups WHERE id=?", (ch_group_id,))
    selected_group_name = cursor.fetchone()[0]

    cursor.execute('SELECT name FROM nameplates WHERE ch_group_id=?', (ch_group_id,))
    existing_names = set([row[0] for row in cursor.fetchall()])

    def get_entries(group_id):
        cursor.execute('''
            SELECT n.sl_no, n.name, p.default_size, n.qty, n.repeater
            FROM nameplates n
            JOIN plate_types p ON n.type_id = p.id
            WHERE n.ch_group_id=? 
        ''', (group_id,))
        return cursor.fetchall()

    entries = get_entries(ch_group_id)
    common_entries = get_entries(common_id)
    filtered_common = []

    for e in common_entries:
        sl_no, name, size, qty, repeater = e
        if name in existing_names:
            continue
        if repeater == 2 and selected_group_name == 'SFCB':
            continue
        elif repeater == 1 or (repeater == 2 and selected_group_name != 'SFCB'):
            for suffix in repeater_map.get(selected_group_name, ['']):
                filtered_common.append((sl_no, f"{name} {suffix}".strip(), size, qty))
        else:
            filtered_common.append((sl_no, name, size, qty))

    all_entries = entries + filtered_common
    ring_entries = [e[:4] for e in all_entries if 'Φ' in e[2]]
    rect_entries = [e[:4] for e in all_entries if 'x' in e[2]]

    conn.close()
    return ring_entries, rect_entries



# ------------------- Main App -------------------
class NameplateApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Nameplate Ordering App")
        self.setMinimumSize(800, 600)
        self.setup_ui()

    def setup_ui(self):
        # --- Main layout ---
        main_layout = QVBoxLayout()
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # --- Inputs: Customer, Job, Charger Group ---
        form_layout = QHBoxLayout()
        form_layout.setSpacing(10)

        lbl_customer = QLabel("Customer Name:")
        self.txt_customer = QLineEdit()
        self.txt_customer.setText("Default Customer")
        lbl_job = QLabel("Job Number:")
        self.txt_job = QLineEdit()
        self.txt_job.setText("1234")
        lbl_group = QLabel("Charger Group:")
        self.cmb_group = QComboBox()

        form_layout.addWidget(lbl_customer)
        form_layout.addWidget(self.txt_customer)
        form_layout.addWidget(lbl_job)
        form_layout.addWidget(self.txt_job)
        form_layout.addWidget(lbl_group)
        form_layout.addWidget(self.cmb_group)

        main_layout.addLayout(form_layout)

        # --- Special checkbox ---
        self.chk_special = QCheckBox("Include Special Nameplates")
        self.chk_special.setStyleSheet("font-weight: bold;")
        main_layout.addWidget(self.chk_special, alignment=Qt.AlignmentFlag.AlignLeft)

        # --- Generate button ---
        self.btn_generate = QPushButton("📝 Generate Nameplate")
        self.btn_generate.clicked.connect(self.generate_nameplate)
        self.btn_generate.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 20px;
                border: 2px solid #388E3C;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #45A049;
            }
            QPushButton:pressed {
                background-color: #2E7D32;
            }
        """)
        main_layout.addWidget(self.btn_generate, alignment=Qt.AlignmentFlag.AlignLeft)

        # --- Export buttons ---
        export_layout = QHBoxLayout()
        export_layout.setSpacing(10)

        self.btn_export = QPushButton("📊 Export to Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_export_pdf = QPushButton("📄 Export to PDF")
        self.btn_export_pdf.clicked.connect(self.export_to_pdf)

        for btn in [self.btn_export, self.btn_export_pdf]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    font-weight: bold;
                    padding: 8px 16px;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #1976D2;
                }
                QPushButton:pressed {
                    background-color: #0D47A1;
                }
            """)

        export_layout.addWidget(self.btn_export)
        export_layout.addWidget(self.btn_export_pdf)
        main_layout.addLayout(export_layout)

        # --- Bulk Quantity Edit ---
        bulk_layout = QHBoxLayout()
        bulk_layout.setSpacing(10)

        lbl_bulk_qty = QLabel("🔢 Set Qty for all entries:")
        self.spin_bulk_qty = QSpinBox()
        self.spin_bulk_qty.setMinimum(1)
        self.spin_bulk_qty.setMaximum(9999)
        self.spin_bulk_qty.setValue(1)

        self.btn_bulk_qty = QPushButton("⚡ Update All Quantities")
        self.btn_bulk_qty.clicked.connect(self.bulk_update_qty)
        self.btn_bulk_qty.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #FB8C00;
            }
            QPushButton:pressed {
                background-color: #EF6C00;
            }
        """)

        bulk_layout.addWidget(lbl_bulk_qty)
        bulk_layout.addWidget(self.spin_bulk_qty)
        bulk_layout.addWidget(self.btn_bulk_qty)
        main_layout.addLayout(bulk_layout)

        # --- Clear All Entries ---
        self.btn_clear_all = QPushButton("🗑️ Clear All Entries")
        self.btn_clear_all.clicked.connect(self.clear_all_entries)
        self.btn_clear_all.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #E53935;
            }
            QPushButton:pressed {
                background-color: #C62828;
            }
        """)
        main_layout.addWidget(self.btn_clear_all, alignment=Qt.AlignmentFlag.AlignLeft)

        # --- Heading Label ---
        self.lbl_heading = QLabel("")
        self.lbl_heading.setStyleSheet("font-weight: bold; font-size: 16px; color: #333333;")
        main_layout.addWidget(self.lbl_heading)

        # --- Result Table ---
        self.tbl_result = QTableWidget(0, 4)
        self.tbl_result.setHorizontalHeaderLabels(['SL No', 'Nameplate Name', 'Cutout/Size', 'Qty'])
        self.tbl_result.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_result.setAlternatingRowColors(True)
        # REMOVED: self.tbl_result.setStyleSheet("alternate-background-color: #f9f9f9; background-color: #ffffff;")
        # The table will now inherit its alternating and background colors 
        # from the global application stylesheet (DARK_STYLE or LIGHT_STYLE).
        main_layout.addWidget(self.tbl_result)

        # --- Add/Remove Buttons ---
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.btn_add_rect = QPushButton("➕ Add Rectangular Entry")
        self.btn_add_ring = QPushButton("➕ Add Ring Entry")
        self.btn_remove = QPushButton("❌ Remove Selected Entry")

        for btn in [self.btn_add_rect, self.btn_add_ring, self.btn_remove]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #9C27B0;
                    color: white;
                    font-weight: bold;
                    padding: 6px 14px;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #7B1FA2;
                }
                QPushButton:pressed {
                    background-color: #6A1B9A;
                }
            """)

        self.btn_add_rect.clicked.connect(lambda: self.add_custom_entry("RECT"))
        self.btn_add_ring.clicked.connect(lambda: self.add_custom_entry("RING"))
        self.btn_remove.clicked.connect(self.remove_selected_entry)

        btn_layout.addWidget(self.btn_add_rect)
        btn_layout.addWidget(self.btn_add_ring)
        btn_layout.addWidget(self.btn_remove)
        main_layout.addLayout(btn_layout)

        # --- Set main layout ---
        self.setLayout(main_layout)

        # --- Load Charger Groups ---
        self.load_groups()


    def load_groups(self):
        # Load charger groups excluding COMMON and SPECIAL
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT group_name FROM ch_groups WHERE group_name NOT IN ('COMMON','SPECIAL')")
        groups = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.cmb_group.addItems(groups)

    def generate_nameplate(self):
        customer = self.txt_customer.text().strip()
        job_no = self.txt_job.text().strip()
        ch_group_name = self.cmb_group.currentText()

        if not customer or not job_no or not ch_group_name:
            QMessageBox.critical(self, "Input Error", "Please fill all fields.")
            return

        # Fetch ch_group_id
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM ch_groups WHERE group_name=?', (ch_group_name,))
        result = cursor.fetchone()
        conn.close()
        if not result:
            QMessageBox.critical(self, "Error", f"Charger group '{ch_group_name}' not found!")
            return
        ch_group_id = result[0]

        # Fetch entries from DB
        ring_entries, rect_entries = fetch_nameplates(ch_group_id)

        # Include SPECIAL if checkbox checked
        if self.chk_special.isChecked():
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM ch_groups WHERE group_name='SPECIAL'")
            special_group = cursor.fetchone()
            conn.close()
            if special_group:
                special_id = special_group[0]
                special_ring, special_rect = fetch_nameplates(special_id)
                ring_entries.extend(special_ring)
                rect_entries.extend(special_rect)

        # Clear table
        self.tbl_result.setRowCount(0)

        # Set heading
        self.lbl_heading.setText(f"LIVELINE NAME-PLATE-{job_no}-{ch_group_name} - {customer}")

        # Track added entries to avoid duplicates
        added_entries = set()

        # --- Insert Rectangular Section ---
        if rect_entries:
            self.add_table_row(("--- RECTANGULAR TYPE ---", "", "", ""))
            for entry in rect_entries:
                self.add_table_row(entry, section="RECT")

        # --- Insert Ring Section ---
        if ring_entries:
            self.add_table_row(("--- RING TYPE ---", "", "", ""))
            for entry in ring_entries:
                self.add_table_row(entry, section="RING")


    def add_table_row(self, row_data, section=None):
        """
        row_data: tuple of (sl_no, name, cutout, qty) or section header
        section: 'RECT' or 'RING' if this row belongs to a section
        """
        row = self.tbl_result.rowCount()
        self.tbl_result.insertRow(row)
        
        for col, value in enumerate(row_data):
            # SL No handling
            if col == 0 and section and not str(value).startswith("---"):
                value = self.get_next_sl_no(section)
            item = QTableWidgetItem(str(value))
            if str(value).startswith("---"):
                font = QFont()
                font.setBold(True)
                item.setFont(font)
                item.setBackground(Qt.GlobalColor.lightGray)
            self.tbl_result.setItem(row, col, item)

    def get_next_sl_no(self, section):
        """
        Counts rows in the table for the given section and returns next incremental SL No.
        """
        count = 0
        in_section = False
        header_text = "--- RECTANGULAR TYPE ---" if section == "RECT" else "--- RING TYPE ---"
        
        for row in range(self.tbl_result.rowCount()):
            item = self.tbl_result.item(row, 0)
            if not item:
                continue
            text = item.text()
            if text == header_text:
                in_section = True
                continue
            if text.startswith("---") and text != header_text:
                in_section = False
            if in_section:
                count += 1
        return count + 1

    def add_custom_entry(self, type_section='RECT'):
        """
        Adds a new empty row in the specified section (RECT or RING) at the BOTTOM of the section.
        Creates the section if it doesn't exist yet.
        """
        header_text = "--- RECTANGULAR TYPE ---" if type_section == 'RECT' else "--- RING TYPE ---"
        section_found = False
        last_row_in_section = -1

        # Find section header
        for row in range(self.tbl_result.rowCount()):
            item = self.tbl_result.item(row, 0)
            if item and item.text() == header_text:
                section_found = True
                last_row_in_section = row
                # Walk forward to find the last row in this section
                temp_row = row + 1
                while temp_row < self.tbl_result.rowCount():
                    next_item = self.tbl_result.item(temp_row, 0)
                    if next_item and next_item.text().startswith("---"):
                        break  # next section reached
                    last_row_in_section = temp_row
                    temp_row += 1
                break

        # If section not found, create it at the end
        if not section_found:
            last_row_in_section = self.tbl_result.rowCount()
            self.tbl_result.insertRow(last_row_in_section)
            header_item = QTableWidgetItem(header_text)
            header_font = QFont()
            header_font.setBold(True)
            header_item.setFont(header_font)
            header_item.setBackground(Qt.GlobalColor.lightGray)
            self.tbl_result.setItem(last_row_in_section, 0, header_item)
            last_row_in_section += 1  # new row will be inserted after header

        # Insert new row **after the last row in the section**
        insert_row = last_row_in_section + 1
        self.tbl_result.insertRow(insert_row)

        # Fill row data
        for col in range(4):
            value = ""
            if col == 0:
                value = self.get_next_sl_no(type_section)
            self.tbl_result.setItem(insert_row, col, QTableWidgetItem(str(value)))

        self.tbl_result.selectRow(insert_row)


            
    def remove_selected_entry(self):
        selected_rows = set(item.row() for item in self.tbl_result.selectedItems())

        if not selected_rows:
            QMessageBox.warning(self, "Remove Entry", "No entries selected.")
            return

        removed_count = 0
        for row in sorted(selected_rows, reverse=True):
            first_cell = self.tbl_result.item(row, 0)
            # Skip section headers
            if first_cell and first_cell.text().startswith("---"):
                continue
            self.tbl_result.removeRow(row)
            removed_count += 1

        QMessageBox.information(self, "Remove Entry", f"Removed {removed_count} selected entries.")


    def export_to_excel(self):
        if self.tbl_result.rowCount() == 0:
            QMessageBox.warning(self, "Export Error", "No data to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        MAX_LEN = 30  # wrap threshold for Nameplate Name

        wb = Workbook()
        ws = wb.active
        ws.title = "Nameplates"

        # Styles
        center_align = Alignment(horizontal="center")
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        
        rect_fill = PatternFill(start_color='FFCCE5FF', end_color='FFCCE5FF', fill_type='solid')  # light blue
        ring_fill = PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')  # light orange

        # Main header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        header_cell = ws.cell(row=1, column=1)
        header_cell.value = self.lbl_heading.text()
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = center_align

        current_row = 3
        current_section = None  # Track current section for coloring

        for r in range(self.tbl_result.rowCount()):
            row_values = [self.tbl_result.item(r, c).text() if self.tbl_result.item(r, c) else "" 
                        for c in range(self.tbl_result.columnCount())]

            # Section heading
            if row_values[0].startswith("---"):
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws.cell(row=current_row, column=1, value=row_values[0])
                cell.font = Font(bold=True, color="080707")
                cell.alignment = center_align
                cell.fill = rect_fill if "RECTANGULAR" in row_values[0] else ring_fill
                current_section = "RECT" if "RECTANGULAR" in row_values[0] else "RING"
                current_row += 1

                # Column headers
                headers = ['SL No', 'Nameplate Name', 'Cutout/Size', 'Qty']
                for col, header in enumerate(headers, start=1):
                    hcell = ws.cell(row=current_row, column=col, value=header)
                    hcell.font = Font(bold=True)
                    hcell.alignment = center_align
                    hcell.border = thin_border
                    hcell.fill = rect_fill if current_section=="RECT" else ring_fill
                current_row += 1
            else:
                for col, value in enumerate(row_values, start=1):
                    # Wrap Nameplate Name column
                    if col == 2 and len(value) > MAX_LEN:
                        wrapped = ""
                        while len(value) > MAX_LEN:
                            split_at = value.rfind(' ', 0, MAX_LEN)
                            if split_at == -1:
                                split_at = MAX_LEN
                            wrapped += value[:split_at] + "\n"
                            value = value[split_at:].lstrip()
                        wrapped += value
                        value = wrapped

                    cell = ws.cell(row=current_row, column=col, value=value)
                    # Wrap text for Nameplate Name
                    if col == 2:
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    else:
                        cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8

        try:
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Excel file saved at:\n{file_path}")
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            elif os.name == 'posix':  # macOS/Linux
                subprocess.call(['open' if sys.platform=='darwin' else 'xdg-open', file_path])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel file:\n{e}")


    def export_to_pdf(self):
        if self.tbl_result.rowCount() == 0:
            QMessageBox.warning(self, "Export Error", "No data to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF File", "", "PDF Files (*.pdf)")
        if not file_path:
            return

        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        MAX_LEN = 30
        wrap_style = ParagraphStyle(
            name='WrapStyle',
            fontName='Helvetica',
            fontSize=10,
            leading=12,
            alignment=1,  # center
        )

        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Header
        header = Paragraph(self.lbl_heading.text(), styles['Title'])
        elements.append(header)
        elements.append(Spacer(1, 12))

        # Build table data with wrapping
        data = []
        for r in range(self.tbl_result.rowCount()):
            row_values = []
            for c in range(self.tbl_result.columnCount()):
                val = self.tbl_result.item(r, c).text() if self.tbl_result.item(r, c) else ""
                # Wrap Nameplate Name column
                if c == 1 and len(val) > MAX_LEN:
                    val = Paragraph(val, wrap_style)
                row_values.append(val)
            data.append(row_values)

        # Table style
        table = Table(data, colWidths=[50, 250, 80, 50])
        style = TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ])

        # Highlight section headers
        for i, row in enumerate(data):
            first_cell = row[0]
            if isinstance(first_cell, str) and first_cell.startswith("---"):
                style.add('BACKGROUND', (0,i), (-1,i), colors.lightgrey)
                style.add('SPAN', (0,i), (-1,i))
                style.add('ALIGN', (0,i), (-1,i), 'CENTER')
                style.add('FONTNAME', (0,i), (-1,i), 'Helvetica-Bold')

        table.setStyle(style)
        elements.append(table)

        try:
            doc.build(elements)
            QMessageBox.information(self, "Success", f"PDF file saved at:\n{file_path}")
            if os.name == 'nt':
                os.startfile(file_path)
            elif os.name == 'posix':
                subprocess.call(['open' if sys.platform=='darwin' else 'xdg-open', file_path])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save PDF file:\n{e}")

    def bulk_update_qty(self):
        new_qty = self.spin_bulk_qty.value()

        updated_rows = 0
        for row in range(self.tbl_result.rowCount()):
            # Skip section headers
            first_cell = self.tbl_result.item(row, 0)
            if first_cell and first_cell.text().startswith("---"):
                continue

            # Update Qty column (last column, index 3)
            qty_item = self.tbl_result.item(row, 3)
            if qty_item:
                qty_item.setText(str(new_qty))
                updated_rows += 1

        QMessageBox.information(self, "Bulk Quantity Update", f"Updated Qty for {updated_rows} entries.")

    def clear_all_entries(self):
        removed_rows = 0
        # Remove only non-header rows
        for row in reversed(range(self.tbl_result.rowCount())):
            first_cell = self.tbl_result.item(row, 0)
            if first_cell and not first_cell.text().startswith("---"):
                self.tbl_result.removeRow(row)
                removed_rows += 1

        QMessageBox.information(self, "Clear Entries", f"Cleared {removed_rows} entries.")
# ------------------- Run App -------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('windows11')
    window = NameplateApp()
    window.show()
    sys.exit(app.exec())
